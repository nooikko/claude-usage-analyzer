"""Excel (.xlsx) export via openpyxl."""

from __future__ import annotations

import logging

from ..utils import cache_hit_rate_float, format_tokens

logger = logging.getLogger(__name__)


def _ensure_openpyxl():
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        raise SystemExit(
            "Excel export requires openpyxl.  Install it with:\n"
            "  pip install claude-usage-analyzer[excel]\n"
            "  # or: pip install openpyxl"
        )


def _add_sheet(wb, title, headers, rows):
    ws = wb.create_sheet(title=title)
    ws.append(headers)
    for row in rows:
        ws.append(list(row))
    # Auto-width (rough)
    for col_idx, header in enumerate(headers, 1):
        max_len = len(str(header))
        for row in rows:
            if col_idx - 1 < len(row):
                max_len = max(max_len, len(str(row[col_idx - 1])))
        ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = min(max_len + 3, 40)
    return ws


def write_report(stats: dict, output: str | None, top_n: int = 20):
    """Write a multi-sheet Excel workbook."""
    _ensure_openpyxl()
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    path = output or "claude_usage.xlsx"
    wb = Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="21262d", end_color="21262d", fill_type="solid")

    # --- Sheet 1: By Model ---
    rows = []
    for model, d in sorted(stats["by_model"].items(),
                            key=lambda x: x[1]["cache_read"] + x[1]["cache_create"] + x[1]["input_tokens"],
                            reverse=True):
        total = d["input_tokens"] + d["cache_read"] + d["cache_create"]
        rows.append([
            model, d["input_tokens"], d["cache_read"], d["cache_create"],
            total,
            cache_hit_rate_float(d["cache_read"], d["cache_create"], d["input_tokens"]),
            d["output_tokens"], d["requests"],
        ])
    ws = _add_sheet(wb, "By Model",
                    ["Model", "Uncached", "Cache Read", "Cache Create",
                     "Total Input", "Hit Rate", "Output", "Requests"], rows)

    # --- Sheet 2: By Project ---
    rows = []
    for proj, d in sorted(stats["by_project"].items(),
                           key=lambda x: x[1]["output_tokens"], reverse=True)[:top_n]:
        rows.append([
            proj, d["input_tokens"], d["cache_read"], d["cache_create"],
            cache_hit_rate_float(d["cache_read"], d["cache_create"], d["input_tokens"]),
            d["output_tokens"], d["requests"], d["sessions"], d["tool_calls"],
        ])
    _add_sheet(wb, "By Project",
               ["Project", "Uncached", "Cache Read", "Cache Create",
                "Hit Rate", "Output", "Requests", "Sessions", "Tool Calls"], rows)

    # --- Sheet 3: By Project + Model ---
    rows = []
    for pm_key, d in sorted(stats["by_project_model"].items(),
                              key=lambda x: x[1]["output_tokens"], reverse=True):
        proj, model = pm_key.split("|", 1)
        rows.append([
            proj, model, d["input_tokens"], d["cache_read"], d["cache_create"],
            cache_hit_rate_float(d["cache_read"], d["cache_create"], d["input_tokens"]),
            d["output_tokens"], d["requests"],
        ])
    _add_sheet(wb, "Project + Model",
               ["Project", "Model", "Uncached", "Cache Read", "Cache Create",
                "Hit Rate", "Output", "Requests"], rows)

    # --- Sheet 4: Daily Cache ---
    rows = []
    for dm_key, d in sorted(stats["cache_by_date"].items()):
        date, model = dm_key.split("|", 1)
        if date == "unknown":
            continue
        rows.append([
            date, model, d["input_tokens"], d["cache_read"], d["cache_create"],
            cache_hit_rate_float(d["cache_read"], d["cache_create"], d["input_tokens"]),
            d["requests"],
        ])
    _add_sheet(wb, "Daily Cache",
               ["Date", "Model", "Uncached", "Cache Read", "Cache Create",
                "Hit Rate", "Requests"], rows)

    # --- Sheet 5: Tool Result Cost ---
    rows = []
    for tool, d in sorted(stats["tool_result_cost"].items(),
                           key=lambda x: x[1]["total_chars"], reverse=True):
        avg = d["total_chars"] // d["count"] if d["count"] else 0
        top_proj = max(d["by_project"].items(), key=lambda x: x[1])[0] if d["by_project"] else ""
        rows.append([
            tool, d["count"], d["total_chars"],
            d["total_chars"] // 4, avg // 4, d["max_single"] // 4, top_proj,
        ])
    _add_sheet(wb, "Tool Cost",
               ["Tool", "Calls", "Total Chars", "Est Tokens",
                "Avg Tok/Call", "Max Single Tok", "Top Project"], rows)

    # --- Sheet 6: Subagents ---
    rows = []
    for sa, d in sorted(stats["by_subagent_type"].items(),
                         key=lambda x: x[1]["count"], reverse=True):
        rk = f"result|{sa}"
        rd = stats["subagent_cost"].get(rk, {})
        rt = rd.get("input_tokens", 0)
        top_proj = max(d["by_project"].items(), key=lambda x: x[1])[0] if d["by_project"] else ""
        rows.append([
            sa, d["count"], rt,
            rt // d["count"] if d["count"] else 0, top_proj,
        ])
    _add_sheet(wb, "Subagents",
               ["Type", "Dispatches", "Result Tokens", "Avg Result", "Top Project"], rows)

    # --- Sheet 7: Sessions ---
    items = [(sid, s) for sid, s in stats["sessions"].items()
             if s["input_tokens"] + s["output_tokens"] > 0]
    items.sort(key=lambda x: x[1]["input_tokens"] + x[1]["output_tokens"], reverse=True)
    rows = []
    for sid, s in items[:200]:
        rows.append([
            sid, s["project"], s["input_tokens"], s["output_tokens"],
            s["cache_read"], s["cache_create"],
            cache_hit_rate_float(s["cache_read"], s["cache_create"], s["input_tokens"]),
            s["tool_calls"],
            ", ".join(sorted(s["models"])),
            s.get("start", "")[:19] if s.get("start") else "",
        ])
    _add_sheet(wb, "Sessions",
               ["Session ID", "Project", "Uncached", "Output",
                "Cache Read", "Cache Create", "Hit Rate", "Tools",
                "Models", "Start"], rows)

    # Style headers
    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
        ws.freeze_panes = "A2"

    wb.save(path)
    logger.info("Excel workbook written to %s", path)
    print(f"Excel workbook written to {path}")
